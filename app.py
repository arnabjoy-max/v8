import streamlit as st
import openpyxl
from datetime import datetime, date, timedelta
import io, smtplib, os
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

# ─── Page Config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="FANUC FIR Generator",
    page_icon="🤖",
    layout="centered",
    initial_sidebar_state="collapsed"
)

# ══════════════════════════════════════════════════════════════════════════════
# AD CONFIGURATION  —  flip SHOW_ADS = True and fill IDs to enable
# ══════════════════════════════════════════════════════════════════════════════
SHOW_ADS             = False
ADSENSE_PUBLISHER_ID = "ca-pub-XXXXXXXXXXXXXXXX"
ADSENSE_SLOT_TOP     = "1111111111"
ADSENSE_SLOT_MID     = "2222222222"
ADSENSE_SLOT_BOTTOM  = "3333333333"

def inject_adsense():
    if not SHOW_ADS: return
    st.markdown(
        f'<script async src="https://pagead2.googlesyndication.com/pagead/js/adsbygoogle.js'
        f'?client={ADSENSE_PUBLISHER_ID}" crossorigin="anonymous"></script>',
        unsafe_allow_html=True)

def ad_slot(slot_id, height="90px"):
    if not SHOW_ADS: return
    st.markdown(f"""
    <div style="text-align:center;margin:10px 0;">
      <ins class="adsbygoogle" style="display:block;width:100%;height:{height}"
           data-ad-client="{ADSENSE_PUBLISHER_ID}" data-ad-slot="{slot_id}"
           data-ad-format="auto" data-full-width-responsive="true"></ins>
      <script>(adsbygoogle = window.adsbygoogle || []).push({{}});</script>
    </div>""", unsafe_allow_html=True)

inject_adsense()

# ── Auto-scroll to top on every page render ───────────────────────────────────
# Uses a unique key each render so Streamlit doesn't cache the script away
st.markdown(
    f'<script>window.parent.document.querySelector("section.main").scrollTo({{top:0,behavior:"instant"}});</script>',
    unsafe_allow_html=True)

# ─── FANUC Logo SVG ───────────────────────────────────────────────────────────
FANUC_LOGO_SVG = """
<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 200 60" width="110" height="33">
  <rect width="200" height="60" fill="#FFD700"/>
  <text x="100" y="46" font-family="Arial Black, Arial" font-weight="900"
        font-size="42" fill="#CC0000" text-anchor="middle" letter-spacing="-1">FANUC</text>
</svg>"""

# ══════════════════════════════════════════════════════════════════════════════
# CSS — V4 LIGHT THEME, FORCED WHITE, PLACEHOLDER INVISIBLE
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');

/* ── Force white background on EVERYTHING — overrides any phone/browser dark mode ── */
:root {
  --bg:       #F0F2F5;
  --white:    #FFFFFF;
  --border:   #E0E3E8;
  --txt:      #1A1A2E;
  --txt2:     #6B7280;
  --yellow:   #FFD700;
  --red:      #CC0000;
  --radius:   14px;
  --shadow:   0 2px 12px rgba(0,0,0,0.07);
  color-scheme: light !important;
}

html, body,
[class*="css"],
[data-testid="stAppViewContainer"],
[data-testid="stMain"],
section[data-testid="stMain"],
.stMainBlockContainer,
.block-container {
  font-family: 'Inter', sans-serif !important;
  background:       #F0F2F5 !important;
  background-color: #F0F2F5 !important;
  color: #1A1A2E !important;
  color-scheme: light !important;
}

/* ── Wipe every Streamlit chrome bar ── */
#MainMenu, footer, header                { visibility:hidden !important; display:none !important; }
[data-testid="stHeader"]                 { display:none !important; height:0 !important; }
[data-testid="stToolbar"]                { display:none !important; height:0 !important; }
[data-testid="stDecoration"]             { display:none !important; height:0 !important; }
[data-testid="stStatusWidget"]           { display:none !important; }
.stDeployButton                          { display:none !important; }
[data-testid="stAppViewBlockContainer"]  { padding-top:0 !important; margin-top:0 !important; }
[data-testid="stAppViewBlockContainer"] > div:first-child { padding-top:0 !important; }
[data-testid="stMain"] > div:first-child { padding-top:0 !important; }
section[data-testid="stMain"]            { padding-top:0 !important; margin-top:0 !important; }
.stMainBlockContainer                    { padding-top:0 !important; }

/* ── Layout ── */
.block-container {
  padding: 0 0 90px !important;
  max-width: 480px !important;
  margin: 0 auto !important;
}

/* ── Top bar ── */
.topbar {
  background: #1A1A2E;
  padding: 0 16px;
  display: flex;
  align-items: center;
  justify-content: space-between;
  height: 58px;
  position: sticky;
  top: 0;
  z-index: 9999;
  box-shadow: 0 2px 8px rgba(0,0,0,0.25);
}

/* ── Step pills ── */
.steps-wrap {
  background: #FFFFFF;
  border-bottom: 1px solid #E0E3E8;
  padding: 10px 14px;
}
.steps { display:flex; gap:6px; overflow-x:auto; scrollbar-width:none; }
.steps::-webkit-scrollbar { display:none; }
.sp {
  flex-shrink:0; border-radius:20px; padding:6px 14px;
  font-size:12px; font-weight:600; letter-spacing:0.5px;
  white-space:nowrap; border:1.5px solid #E0E3E8;
  background:#FFFFFF; color:#6B7280;
}
.sp.on   { background:#1A1A2E; color:#fff; border-color:#1A1A2E; }
.sp.done { background:#F0FFF4; color:#16A34A; border-color:#86EFAC; }

/* ── Page title ── */
.page-title { padding:22px 16px 6px; }
.page-title .icon { font-size:32px; display:block; margin-bottom:6px; }
.page-title h2 {
  font-size:26px; font-weight:700; color:#1A1A2E;
  margin:0 0 2px; line-height:1.2;
}
.page-title p { font-size:13px; color:#6B7280; margin:0; }

/* ── Cards ── */
.card {
  background: #FFFFFF;
  border-radius: var(--radius);
  box-shadow: var(--shadow);
  padding: 20px 16px;
  margin: 12px 12px;
}
.card-label {
  font-size:11px; font-weight:700; color:#6B7280;
  letter-spacing:1.5px; text-transform:uppercase;
  margin-bottom:14px;
  display:flex; align-items:center; gap:8px;
}
.card-label::after { content:''; flex:1; height:1px; background:#E0E3E8; }

/* ── Day cards ── */
.day-card {
  background: #FFFFFF;
  border: 1.5px solid #E0E3E8;
  border-radius: 12px;
  padding: 14px 16px 0 16px;
  margin: 0 12px 2px;
  box-shadow: 0 1px 6px rgba(0,0,0,0.05);
}
.day-date-label {
  font-size:10px; font-weight:700; color:#6B7280;
  letter-spacing:1.5px; text-transform:uppercase; margin-bottom:2px;
}
.day-date-val { font-size:17px; font-weight:700; color:#1A1A2E; margin-bottom:10px; }

/* ── Filename badge ── */
.fname {
  background: #FFFBEB;
  border: 1.5px solid #FFD700;
  border-radius: 10px;
  padding: 12px 14px;
  font-size: 13px; font-weight: 700;
  color: #92400E;
  text-align: center;
  word-break: break-all;
  margin: 12px 0 4px;
}
.fname-label {
  font-size:10px; color:#6B7280; text-align:center;
  letter-spacing:1px; text-transform:uppercase; margin-bottom:4px;
}

/* ── Summary row ── */
.summary-row {
  display:grid; grid-template-columns:repeat(4,1fr);
  gap:8px; margin:12px 12px;
}
.summary-cell {
  background:#FFFFFF; border-radius:10px;
  box-shadow:0 2px 8px rgba(0,0,0,0.06); padding:10px 6px; text-align:center;
}
.summary-cell .val { font-size:18px; font-weight:700; color:#1A1A2E; }
.summary-cell .lbl { font-size:10px; color:#6B7280; letter-spacing:1px; text-transform:uppercase; margin-top:2px; }

/* ── Inputs ── */
input, textarea, select,
.stTextInput > div > div > input,
.stSelectbox > div > div,
[data-baseweb="select"] > div,
.stDateInput > div > div > input,
.stNumberInput > div > div > input,
.stTextArea > div > textarea {
  background:   #F9FAFB !important;
  border: 1.5px solid #E0E3E8 !important;
  border-radius: 10px !important;
  color: #1A1A2E !important;
  font-size: 16px !important;
  min-height: 50px !important;
  box-shadow: none !important;
}
input:focus, textarea:focus,
.stTextInput > div > div > input:focus,
.stTextArea > div > textarea:focus {
  border-color: #1A1A2E !important;
  background: #FFFFFF !important;
  box-shadow: 0 0 0 2px rgba(26,26,46,0.08) !important;
}

/* ── Placeholder — nearly invisible ── */
::placeholder                             { color:#D1D5DB !important; opacity:1 !important; }
::-webkit-input-placeholder               { color:#D1D5DB !important; }
::-moz-placeholder                        { color:#D1D5DB !important; opacity:1 !important; }
input::placeholder, textarea::placeholder { color:#D1D5DB !important; }

/* Dropdown options */
[data-baseweb="popover"] *, ul[role="listbox"] * {
  background: #FFFFFF !important;
  color: #1A1A2E !important;
}
[data-baseweb="option"]:hover { background: #F3F4F6 !important; }

/* Labels */
label, .stTextInput label, .stSelectbox label,
.stDateInput label, .stNumberInput label,
.stTextArea label, .stCheckbox label {
  color: #6B7280 !important;
  font-size: 11px !important;
  font-weight: 700 !important;
  letter-spacing: 1.2px !important;
  text-transform: uppercase !important;
}

/* ── Buttons ── */
.stButton > button {
  width: 100% !important;
  min-height: 52px !important;
  background: #1A1A2E !important;
  color: #fff !important;
  font-family: 'Inter', sans-serif !important;
  font-weight: 700 !important;
  font-size: 15px !important;
  letter-spacing: 1.5px !important;
  border: none !important;
  border-radius: 12px !important;
  margin: 4px 0 !important;
  transition: opacity 0.15s !important;
}
.stButton > button:hover { opacity: 0.88 !important; }

/* Back / secondary — left column button */
div[data-testid="column"]:first-child .stButton > button {
  background: #F3F4F6 !important;
  color: #374151 !important;
  border: 1.5px solid #E0E3E8 !important;
}

/* Delete button — keep it red-ish */
.del-btn .stButton > button {
  background: #FEF2F2 !important;
  color: #DC2626 !important;
  border: 1.5px solid #FECACA !important;
  font-size: 18px !important;
  min-height: 50px !important;
}

/* Download button */
.stDownloadButton > button {
  width: 100% !important;
  min-height: 52px !important;
  background: #FFD700 !important;
  color: #000 !important;
  font-weight: 700 !important;
  font-size: 15px !important;
  letter-spacing: 1.5px !important;
  border: none !important;
  border-radius: 12px !important;
}

/* ── Alerts ── */
.stAlert { border-radius:10px !important; font-size:14px !important; }

/* ── Bottom nav ── */
.bottom-nav { padding: 0 12px; margin-top: 8px; }

/* ── Divider / scrollbar ── */
hr { border-color: #E0E3E8 !important; margin: 14px 0 !important; }
::-webkit-scrollbar { width: 4px; }
::-webkit-scrollbar-thumb { background: #E0E3E8; border-radius: 2px; }
</style>
""", unsafe_allow_html=True)


# ─── Constants ────────────────────────────────────────────────────────────────
REPORT_TYPES = {"Installation (I)":"I", "Production Support (P)":"P", "Other (O)":"O"}

# ─── Helpers ──────────────────────────────────────────────────────────────────
def next_monday(d):
    """Return d if already Monday, else the Monday of that week."""
    return d - timedelta(days=d.weekday())

def generate_filename(project_num, start_date, emp_no, last_name, report_type):
    if not all([project_num, start_date]):
        return "FIR-______-______-___(_).xlsm"
    identifier = (emp_no or "").strip() or (last_name or "").strip() or "___"
    suffix = REPORT_TYPES.get(report_type, "O")
    return f"FIR-{project_num}-{start_date.strftime('%y%m%d')}-{identifier}({suffix}).xlsm"

def populate_fir_excel(data):
    template_path = os.path.join(os.path.dirname(__file__), "FIR_template.xlsm")
    wb  = openpyxl.load_workbook(template_path, keep_vba=True)
    ws  = wb["Field Installation"]
    ws2 = wb["SUMFIR"]

    ws["I2"] = int(data["project_num"]) if data["project_num"].isdigit() else data["project_num"]
    ws["O2"] = data.get("program_process_cell", "")
    ws["D3"] = data.get("customer_contact", "")
    ws["L3"] = data.get("phone", "")
    ws["S3"] = data.get("plant_location", "")
    ws2.cell(row=2, column=48).value = {"Installation (I)":1,"Production Support (P)":2,"Other (O)":3}.get(data["report_type"], 3)
    ws2.cell(row=2, column=49).value = 1 if data.get("exp_rep") else 2
    ws["X5"] = data.get("expense_amount", 0) or 0
    ws["E55"] = data.get("engineer_first_name", "")
    emp = (data.get("emp_no") or "").strip()
    ws["M55"] = emp if emp else (data.get("last_name") or "").strip()
    ws["B56"] = data.get("engineer_print_name", "")

    # day_map: (date_row, hours_row, work_row)
    day_map = [(6,9,6),(12,15,12),(18,21,18),(24,27,24),(30,33,30),(36,39,36),(42,45,42)]
    for i, (dr, hr, wr) in enumerate(day_map):
        d = data["days"][i] if i < len(data["days"]) else {}
        if not d.get("active"):
            continue
        # Write each day's date explicitly — no need to hit "Edit" in Excel
        day_date = d.get("day_date")
        if day_date:
            ws.cell(row=dr, column=1).value = datetime.combine(day_date, datetime.min.time())
        for col, key in [(1,"straight"),(2,"overtime"),(3,"doubletime"),(4,"travel_time"),(5,"working"),(27,"wait")]:
            val = d.get(key)
            ws.cell(row=hr, column=col).value = val if val else None
        ws.cell(row=wr, column=6).value = d.get("description", "")

    ws.cell(row=47, column=6).value = data.get("additional_comments", "")
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()

def send_email_report(smtp_host, smtp_port, smtp_user, smtp_pass, recipients, subject, body, attachment, filename):
    msg = MIMEMultipart()
    msg["From"] = smtp_user
    msg["To"]   = ", ".join(recipients)
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain"))
    part = MIMEBase("application", "octet-stream")
    part.set_payload(attachment)
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{filename}"')
    msg.attach(part)
    with smtplib.SMTP(smtp_host, smtp_port) as s:
        s.starttls(); s.login(smtp_user, smtp_pass)
        s.sendmail(smtp_user, recipients, msg.as_string())


# ─── Session state defaults ───────────────────────────────────────────────────
_day_empty = lambda: {"active":True, "day_date":None, "straight":None, "overtime":None,
                       "doubletime":None, "travel_time":None, "working":None,
                       "wait":None, "description":""}

for k, v in [("step",1),("days",[]),("generated_bytes",None),
             ("generated_filename",""),("additional_comments","")]:
    if k not in st.session_state:
        st.session_state[k] = v

step = st.session_state.step

# ─── TOP BAR ─────────────────────────────────────────────────────────────────
st.markdown(f"""
<div class="topbar">
  <div>{FANUC_LOGO_SVG}</div>
  <div style="color:#8B8FA8;font-size:11px;font-weight:600;letter-spacing:1.5px;">FIELD INSTALLATION REPORT</div>
</div>""", unsafe_allow_html=True)

ad_slot(ADSENSE_SLOT_TOP, "60px")

# ─── Step pills ──────────────────────────────────────────────────────────────
pills = [("1","Project"),("2","Engineer"),("3","Days"),("4","Generate"),("5","Email")]
st.markdown(
    '<div class="steps-wrap"><div class="steps">' +
    "".join(f'<div class="sp {"on" if int(n)==step else ("done" if int(n)<step else "")}">{n} · {lbl}</div>'
            for n, lbl in pills) +
    '</div></div>', unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════
# STEP 1 – PROJECT INFO
# ══════════════════════════════════════════════════════════════════
if step == 1:
    st.markdown("""
    <div class="page-title">
      <span class="icon">🏭</span>
      <h2>Project &amp;<br>Customer Details</h2>
    </div>""", unsafe_allow_html=True)

    st.markdown('<div class="card">', unsafe_allow_html=True)
    project_num          = st.text_input("Project #",                value=st.session_state.get("project_num",""),          placeholder="3011187")
    customer_contact     = st.text_input("Customer Contact Name",    value=st.session_state.get("customer_contact",""),      placeholder="Eric Oliveri")
    plant_location       = st.text_input("Plant / Location",         value=st.session_state.get("plant_location",""),        placeholder="Ford Oakville Body Decking")
    program_process_cell = st.text_input("Program / Process / Cell", value=st.session_state.get("program_process_cell",""), placeholder="Body Shop / Cell 12")
    phone                = st.text_input("Customer Phone #",         value=st.session_state.get("phone",""),                 placeholder="586-206-6284")
    report_type          = st.selectbox("Report Type",               list(REPORT_TYPES.keys()),
                                         index=list(REPORT_TYPES.keys()).index(st.session_state.get("report_type","Other (O)")))
    col_exp, col_rep = st.columns([2,1])
    with col_exp:
        expense_amount = st.number_input("Expense Amount ($)", value=float(st.session_state.get("expense_amount",0)), min_value=0.0, step=0.01, format="%.2f")
    with col_rep:
        exp_rep = st.selectbox("Exp Rep?", ["No","Yes"], index=0 if not st.session_state.get("exp_rep",False) else 1)
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="bottom-nav">', unsafe_allow_html=True)
    if st.button("Continue →", use_container_width=True):
        if not project_num:
            st.error("Project # is required.")
        else:
            st.session_state.update(dict(
                project_num=project_num, plant_location=plant_location,
                report_type=report_type, customer_contact=customer_contact, phone=phone,
                program_process_cell=program_process_cell, expense_amount=expense_amount,
                exp_rep=(exp_rep=="Yes"), step=2))
            st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════
# STEP 2 – ENGINEER DETAILS
# ══════════════════════════════════════════════════════════════════
elif step == 2:
    st.markdown("""
    <div class="page-title">
      <span class="icon">👷</span>
      <h2>Engineer<br>Details</h2>
    </div>""", unsafe_allow_html=True)

    st.markdown('<div class="card">', unsafe_allow_html=True)
    engineer_first_name = st.text_input("First Name",      value=st.session_state.get("engineer_first_name",""), placeholder="Arnab")
    last_name           = st.text_input("Last Name",       value=st.session_state.get("last_name",""),           placeholder="Joy")
    engineer_print_name = st.text_input("Full Print Name", value=st.session_state.get("engineer_print_name",""), placeholder="Arnab Joy")
    emp_no              = st.text_input("Employee # (optional — uses Last Name if blank)",
                                         value=st.session_state.get("emp_no",""), placeholder="260216")

    # ── Monday-locked start date ──────────────────────────────────
    _saved = st.session_state.get("start_date", date.today())
    _default_monday = next_monday(_saved)
    start_date_raw = st.date_input("Week Start Date (Monday)", value=_default_monday)
    # Snap to Monday silently if user somehow picks another day
    start_date = next_monday(start_date_raw)
    if start_date != start_date_raw:
        st.info(f"📅 Snapped to Monday · {start_date.strftime('%B %d, %Y')}")
    st.markdown('</div>', unsafe_allow_html=True)

    fname = generate_filename(
        st.session_state.get("project_num",), start_date,
        emp_no, last_name, st.session_state.get("report_type","Other (O)"))
    st.markdown(f"""
    <div style="padding:0 12px;">
      <div class="fname-label">Auto-generated filename</div>
      <div class="fname">{fname}</div>
    </div>""", unsafe_allow_html=True)

    st.markdown('<div class="bottom-nav" style="margin-top:12px;">', unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    with c1:
        if st.button("← Back", use_container_width=True):
            st.session_state.step = 1; st.rerun()
    with c2:
        if st.button("Continue →", use_container_width=True):
            if not engineer_first_name:
                st.error("First Name is required.")
            elif not emp_no and not last_name:
                st.error("Enter Employee # or Last Name.")
            else:
                st.session_state.update(dict(
                    engineer_first_name=engineer_first_name, last_name=last_name,
                    engineer_print_name=engineer_print_name or f"{engineer_first_name} {last_name}",
                    emp_no=emp_no, start_date=start_date, step=3))
                st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════════════════
# STEP 3 – DAILY WORK LOG  (all 7 days Mon–Sun always shown)
# ══════════════════════════════════════════════════════════════════
elif step == 3:
    sd       = st.session_state.get("start_date", date.today())
    week_end = sd + timedelta(days=6)

    # Always initialise exactly 7 days, one per day of the week
    if len(st.session_state.days) != 7:
        st.session_state.days = [
            {"active": False, "day_date": sd + timedelta(days=i),
             "straight": None, "overtime": None, "doubletime": None,
             "travel_time": None, "working": None, "wait": None, "description": ""}
            for i in range(7)
        ]
    days = st.session_state.days

    st.markdown(f"""
    <div class="page-title">
      <span class="icon">📅</span>
      <h2>Daily Work Log</h2>
      <p>Week of {sd.strftime('%b %d')} – {week_end.strftime('%b %d, %Y')}</p>
    </div>""", unsafe_allow_html=True)

    ad_slot(ADSENSE_SLOT_MID, "70px")

    # ── Totals summary ────────────────────────────────────────────
    ts = sum((d.get("straight") or 0) for d in days if d.get("active"))
    to = sum((d.get("overtime") or 0) for d in days if d.get("active"))
    tt = sum((d.get("travel_time") or 0) for d in days if d.get("active"))
    total = ts + to + tt + sum((d.get("doubletime") or 0) for d in days if d.get("active"))
    st.markdown(f"""
    <div class="summary-row">
      <div class="summary-cell"><div class="val">{ts:.0f}</div><div class="lbl">Straight</div></div>
      <div class="summary-cell"><div class="val">{to:.0f}</div><div class="lbl">OT</div></div>
      <div class="summary-cell"><div class="val">{tt:.0f}</div><div class="lbl">Travel</div></div>
      <div class="summary-cell"><div class="val">{total:.0f}</div><div class="lbl">Total</div></div>
    </div>""", unsafe_allow_html=True)

    # ── All 7 day cards Mon–Sun ───────────────────────────────────
    for i in range(7):
        d        = st.session_state.days[i]
        day_date = sd + timedelta(days=i)
        day_name = day_date.strftime("%A")
        date_str = day_date.strftime("%m/%d/%Y")
        is_weekday = day_date.weekday() < 5

        st.markdown(f"""
        <div class="day-card" style="{'border-color:#1A1A2E;' if is_weekday else 'opacity:0.8;'}">
          <div class="day-date-label">Day {i+1} · {day_name}</div>
          <div class="day-date-val">{date_str}</div>
        </div>""", unsafe_allow_html=True)

        with st.container():
        st.markdown('<div style="margin:0 12px;background:#FFFFFF;border:1.5px solid #E0E3E8;border-top:0;border-radius:0 0 12px 12px;padding:4px 14px 14px;margin-top:-2px;">', unsafe_allow_html=True)
            active = st.checkbox(
                f"Include {day_name} in report",
                value=d.get("active", False),
                key=f"act_{i}")
            st.session_state.days[i]["active"] = active

            if active:
                desc = st.text_area(
                    "Work Performed",
                    value=d.get("description",""),
                    placeholder="Describe tasks performed…",
                    key=f"desc_{i}", height=85)
                st.session_state.days[i]["description"] = desc

                h1, h2, h3, h4, h5 = st.columns(5)
                with h1:
                    v = st.number_input("S",  min_value=0.0, max_value=24.0, value=None, step=0.5, key=f"s_{i}",  format="%.1f")
                    st.session_state.days[i]["straight"] = v
                with h2:
                    v = st.number_input("OT", min_value=0.0, max_value=24.0, value=None, step=0.5, key=f"ot_{i}", format="%.1f")
                    st.session_state.days[i]["overtime"] = v
                with h3:
                    v = st.number_input("DT", min_value=0.0, max_value=24.0, value=None, step=0.5, key=f"dt_{i}", format="%.1f")
                    st.session_state.days[i]["doubletime"] = v
                with h4:
                    v = st.number_input("TT", min_value=0.0, max_value=24.0, value=None, step=0.5, key=f"tt_{i}", format="%.1f")
                    st.session_state.days[i]["travel_time"] = v
                with h5:
                    v = st.number_input("W",  min_value=0.0, max_value=24.0, value=None, step=0.5, key=f"w_{i}",  format="%.1f")
                    st.session_state.days[i]["working"] = v

                #v = st.number_input("Wait", min_value=0.0, max_value=24.0, value=None, step=0.5, key=f"wt_{i}", format="%.1f")
                #st.session_state.days[i]["wait"] = v

            st.markdown('</div>', unsafe_allow_html=True)

    # Additional comments
    st.markdown('<div class="card" style="margin-top:8px;">', unsafe_allow_html=True)
    ac = st.text_area("Additional Comments",
                       value=st.session_state.get("additional_comments",""),
                       placeholder="Any extra notes or follow-up items…", height=80)
    st.session_state["additional_comments"] = ac
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="bottom-nav">', unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    with c1:
        if st.button("← Back",     use_container_width=True): st.session_state.step=2; st.rerun()
    with c2:
        if st.button("Continue →", use_container_width=True): st.session_state.step=4; st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)


# STEP 4 – GENERATE & DOWNLOAD
# ══════════════════════════════════════════════════════════════════
elif step == 4:
    fname = generate_filename(
        st.session_state.get("project_num",""),
        st.session_state.get("start_date", date.today()),
        st.session_state.get("emp_no",""),
        st.session_state.get("last_name",""),
        st.session_state.get("report_type","Other (O)"))

    active_days = len(st.session_state.days)
    total_hrs   = sum(
        (d.get("straight") or 0)+(d.get("overtime") or 0)+
        (d.get("doubletime") or 0)+(d.get("travel_time") or 0)
        for d in st.session_state.days)

    st.markdown("""
    <div class="page-title">
      <span class="icon">⚙️</span>
      <h2>Review &amp;<br>Export Report</h2>
    </div>""", unsafe_allow_html=True)

    st.markdown(f"""
    <div class="card">
      <div class="card-label">Report Summary</div>
      <div style="font-size:14px;color:#374151;line-height:2.2;">
        <b>Project:</b> {st.session_state.get('project_num','—')}<br>
        <b>Plant:</b> {st.session_state.get('plant_location','—')}<br>
        <b>Engineer:</b> {st.session_state.get('engineer_print_name','—')}
          &nbsp;·&nbsp; #{st.session_state.get('emp_no', st.session_state.get('last_name','—'))}<br>
        <b>Days:</b> {active_days} &nbsp;·&nbsp; <b>Total Hours:</b> {total_hrs:.1f}
      </div>
      <div class="fname-label" style="margin-top:14px;">Generated Filename</div>
      <div class="fname">{fname}</div>
    </div>""", unsafe_allow_html=True)

    st.markdown('<div class="bottom-nav">', unsafe_allow_html=True)
    if st.button("⚙️  GENERATE FIR REPORT", use_container_width=True):
        with st.spinner("Building your FIR Excel file…"):
            try:
                days_padded = st.session_state.days + [_day_empty()]*(7-len(st.session_state.days))
                xls = populate_fir_excel({
                    "project_num":          st.session_state.get("project_num",""),
                    "plant_location":       st.session_state.get("plant_location",""),
                    "customer_contact":     st.session_state.get("customer_contact",""),
                    "phone":                st.session_state.get("phone",""),
                    "report_type":          st.session_state.get("report_type","Other (O)"),
                    "expense_amount":       st.session_state.get("expense_amount",0),
                    "exp_rep":              st.session_state.get("exp_rep",False),
                    "program_process_cell": st.session_state.get("program_process_cell",""),
                    "engineer_first_name":  st.session_state.get("engineer_first_name",""),
                    "engineer_print_name":  st.session_state.get("engineer_print_name",""),
                    "emp_no":               st.session_state.get("emp_no",""),
                    "last_name":            st.session_state.get("last_name",""),
                    "start_date":           st.session_state.get("start_date", date.today()),
                    "days":                 days_padded,
                    "additional_comments":  st.session_state.get("additional_comments",""),
                })
                st.session_state.generated_bytes    = xls
                st.session_state.generated_filename = fname
                st.success("✅ Report ready — tap Download below!")
            except Exception as e:
                st.error(f"Error generating report: {e}")

    if st.session_state.generated_bytes:
        st.download_button(
            "⬇️  DOWNLOAD  " + st.session_state.generated_filename,
            data=st.session_state.generated_bytes,
            file_name=st.session_state.generated_filename,
            mime="application/vnd.ms-excel.sheet.macroEnabled.12",
            use_container_width=True)
        ad_slot(ADSENSE_SLOT_BOTTOM, "90px")

    c1, c2 = st.columns(2)
    with c1:
        if st.button("← Back",           use_container_width=True): st.session_state.step=3; st.rerun()
    with c2:
        if st.button("📤 Share Report →", use_container_width=True): st.session_state.step=5; st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════
# STEP 5 – SHARE  (mailto — no passwords, opens phone mail app)
# ══════════════════════════════════════════════════════════════════
elif step == 5:
    import urllib.parse

    st.markdown("""
    <div class="page-title">
      <span class="icon">📤</span>
      <h2>Share Report</h2>
      <p>Opens your phone's mail app — no passwords needed</p>
    </div>""", unsafe_allow_html=True)

    if not st.session_state.generated_bytes:
        st.warning("⚠️ Please generate and download the report first (Step 4).")
        st.markdown("""
        <div class="card">
          <div style="font-size:14px;color:#374151;line-height:2;">
            <b>How to share:</b><br>
            1️⃣ Go back to Step 4 and tap <b>Download</b><br>
            2️⃣ The file saves to your phone<br>
            3️⃣ Return here and tap <b>Open in Mail App</b><br>
            4️⃣ Your mail app opens — attach the file and send
          </div>
        </div>""", unsafe_allow_html=True)
    else:
        proj   = st.session_state.get("project_num", "")
        eng    = st.session_state.get("engineer_print_name", "")
        sd_str = st.session_state.get("start_date", date.today()).strftime("%B %d, %Y")
        fname  = st.session_state.generated_filename

        subject = urllib.parse.quote(f"FIR – Project {proj} – {eng}")
        body    = urllib.parse.quote(
            f"Please find attached the Field Installation Report.\n\n"
            f"Project: {proj}\n"
            f"Engineer: {eng}\n"
            f"Week of: {sd_str}\n"
            f"File: {fname}\n\n"
            f"Sent via FANUC FIR System."
        )

        st.markdown(f"""
        <div class="card">
          <div class="card-label">Ready to Share</div>
          <div style="font-size:13px;color:#374151;word-break:break-all;margin-bottom:14px;">
            📎 {fname}
          </div>
          <div style="font-size:13px;color:#374151;line-height:2.1;">
            ✅ File downloaded to your phone<br>
            👇 Tap below to open your mail app<br>
            📎 Attach the file from your Downloads folder
          </div>
        </div>""", unsafe_allow_html=True)

        recipient = st.text_input("Recipient Email (optional — can fill in mail app)",
                                   placeholder="manager@fanuc.com")
        to_part = urllib.parse.quote(recipient.strip()) if recipient.strip() else ""
        mailto_link = f"mailto:{to_part}?subject={subject}&body={body}"

        st.markdown(f"""
        <div style="padding:4px 0 8px;">
          <a href="{mailto_link}" style="
            display:block; width:100%; text-align:center;
            background:#1A1A2E; color:#fff !important;
            -webkit-text-fill-color:#fff !important;
            font-family:Inter,sans-serif; font-weight:700;
            font-size:16px; letter-spacing:1.5px;
            padding:17px 0; border-radius:12px;
            text-decoration:none;
            box-shadow:0 2px 8px rgba(0,0,0,0.15);">
            ✉️ &nbsp; OPEN IN MAIL APP
          </a>
        </div>
        <div style="text-align:center;font-size:12px;color:#9CA3AF;margin-top:4px;">
          No passwords · No server · Uses your phone's built-in mail
        </div>""", unsafe_allow_html=True)

    st.markdown('<div class="bottom-nav" style="margin-top:10px;">', unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    with c1:
        if st.button("← Back",        use_container_width=True): st.session_state.step=4; st.rerun()
    with c2:
        if st.button("🔄 New Report", use_container_width=True):
            for k in ["project_num","plant_location","report_type","customer_contact","phone",
                      "program_process_cell","expense_amount","exp_rep","engineer_first_name",
                      "last_name","engineer_print_name","emp_no","start_date",
                      "additional_comments","generated_bytes","generated_filename"]:
                st.session_state.pop(k, None)
            st.session_state.days = []
            st.session_state.step = 1
            st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)


# ─── Footer ──────────────────────────────────────────────────────────────────
st.markdown("""
<div style="text-align:center;padding:28px 0 10px;color:#9CA3AF;font-size:10px;letter-spacing:1px;">
  FANUC AMERICA CORPORATION · FIR SYSTEM v8.0<br>
  <span style="color:#D1D5DB;">Not for external distribution</span>
</div>""", unsafe_allow_html=True)
